#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import Workbook,load_workbook


excel_write = '../xlsx/target.xlsx'
excel_save = '../xlsx/save.xlsx'


def openpyxl_read():
    wb2 = load_workbook(excel_save)
    print(wb2.get_sheet_names())

    # 取第一张表
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])

def openpyxl_write():
    # 创建一个工作簿
    wb = Workbook()

    # 至少需要一个sheet
    ws = wb.active

    # 如果要指定sheet的序号 可以通过下述方法创建sheet
    # ws1 = wb.create_sheet("Mysheet")  # insert at the end (default)
    # ws2 = wb.create_sheet("Mysheet", 0)  # insert at first position

    # 修改sheet的名称
    ws.title = 'name test'

    # 修改tab的颜色
    ws.sheet_properties.tabColor = "1072BA"

    # 当一个sheet有个名字之后，可以以类似处理字典的方式获得sheet
    # ws3 = wb['name test']

    # 显示sheet名列表
    # print(wb.sheetnames) # ['Sheet2', 'New Title', 'Sheet1']

    # 显示表名，表行数，表列数
    # print
    # "Work Sheet Titile:", ws.title
    # print
    # "Work Sheet Rows:", ws.get_highest_row()
    # print
    # "Work Sheet Cols:", ws.get_highest_column()

    # 拷贝单个sheet
    # Only cells and styles can be copied. You cannot copy worksheets between workbooks.
    # source = wb.active
    # target = wb.copy_worksheet(source)

    # 访问单个单元
    ws['A2'] = 17 # 方法1
    # d = ws.cell(row=1,column=1,value=11) #方法2 行列计数从1开始
    # When a worksheet is created in memory, it contains no cells.
    # They are created when first accessed.

    # 获得单个单元的值
    w1 = ws.cell(row=1, column=2).value
    w2 = ws.cell(row=2, column=2).value

    # 访问多个单元
    cell_range = ws['A1':'C2'] # 获得A1 B1 C1 A2 B2 C2 6个单元格

    # 在访问一个单元或者一个区域时会创建一个从A1到当前访问最大距离的区域
    # 之后对行列访问时，是在范围创建区域内的行列内容，而不是无限范围。
    colC = ws['C'] # 获得C列所有被创建出来的单元格
    col_range = ws['C:D'] # 获得C D两列被创建出来的单元格
    row10 = ws[10]        # 获得第10行被创建出来的单元格
    row_range = ws[5:10] # 获得5-10行被创建的单元格
    # ws.rows # 行访问
    # ws.columns # 列访问

    # This operation will overwrite existing files without warning.
    wb.save(excel_write)


if __name__ == '__main__':
    openpyxl_write()
